import os
import sys
import pulp
import multiprocessing
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import csv
from openpyxl import Workbook  # per export Excel


# ============================================
# FIX PERMESSI CBC (pulp) SU MACOS
# ============================================

def ensure_cbc_executable():
    """
    Su macOS, quando pulp è bundle-ato dentro una .app PyInstaller,
    può capitare che il binario cbc perda il bit di esecuzione.

    Qui cerchiamo il binario CBC esattamente nella struttura che hai verificato:

        TeamOptimizer.app/Contents/Frameworks/pulp/solverdir/cbc/osx/i64/cbc

    e come fallback usiamo anche pulp.__file__ per trovare solverdir.
    """
    if sys.platform != "darwin":
        return

    try:
        candidates = []

        # sys.executable = .../TeamOptimizer.app/Contents/MacOS/TeamOptimizer
        exe_path = sys.executable
        contents_dir = os.path.abspath(os.path.join(exe_path, "..", ".."))   # .../Contents
        frameworks_dir = os.path.join(contents_dir, "Frameworks")
        pulp_from_frameworks = os.path.join(frameworks_dir, "pulp")

        candidates.append(
            os.path.join(pulp_from_frameworks, "solverdir", "cbc", "osx", "i64", "cbc")
        )
        # se un domani ci fosse una build arm64 separata:
        candidates.append(
            os.path.join(pulp_from_frameworks, "solverdir", "cbc", "osx", "arm64", "cbc")
        )

        # 2) Fallback: derivato da pulp.__file__ (per sicurezza)
        pulp_dir = os.path.dirname(pulp.__file__)
        if os.path.basename(pulp_dir) == "apis":
            pulp_dir = os.path.dirname(pulp_dir)

        candidates.append(
            os.path.join(pulp_dir, "solverdir", "cbc", "osx", "i64", "cbc")
        )
        candidates.append(
            os.path.join(pulp_dir, "solverdir", "cbc", "osx", "arm64", "cbc")
        )

        # Deduplica i percorsi
        seen = set()
        unique_candidates = []
        for c in candidates:
            if c not in seen:
                seen.add(c)
                unique_candidates.append(c)

        # Applica chmod +x a tutti quelli che esistono
        for path in unique_candidates:
            if os.path.exists(path):
                st = os.stat(path)
                os.chmod(path, st.st_mode | 0o111)
    except Exception:
        # Non blocchiamo l'app se qualcosa va storto, ci prova e basta
        pass


# ============================================
# UTILS FORMATTING
# ============================================

def euro_fmt(value):
    """Formatta un numero in stile 1.234,56 €"""
    if value is None:
        return "-"
    s = f"{value:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return s + " €"


def pct_fmt(value):
    """value decimale -> '56,48%'"""
    if value is None:
        return "-"
    return f"{value * 100:.2f}".replace(".", ",") + "%"


# ============================================
# PARSING VALORI SENZA % (UI gestisce il suffisso)
# ============================================

def parse_percent_no_symbol(text):
    text = text.strip()
    if text == "":
        return 0.0
    text = text.replace(",", ".")
    return float(text) / 100.0


# ============================================
# PATH RISORSE (CSV accanto all'exe / .app)
# ============================================

def resource_path(relative_path: str) -> str:
    """
    Restituisce il path del file:
    - in sviluppo: cartella del sorgente
    - in eseguibile Windows: cartella del .exe
    - in eseguibile macOS: cartella che CONTIENE la .app
      (così il CSV sta accanto alla .app, non dentro)
    """
    if getattr(sys, 'frozen', False):
        exe_path = sys.executable

        if sys.platform == "darwin":
            # sys.executable = .../TeamOptimizer.app/Contents/MacOS/TeamOptimizer
            # Risalgo di 4 livelli per arrivare alla cartella che contiene la .app:
            # MacOS -> Contents -> TeamOptimizer.app -> <cartella>
            base_path = os.path.dirname(
                os.path.dirname(
                    os.path.dirname(
                        os.path.dirname(exe_path)
                    )
                )
            )
        else:
            # Windows (e altri): cartella dell'exe
            base_path = os.path.dirname(exe_path)
    else:
        # Esecuzione da sorgente (Python normale)
        base_path = os.path.abspath(os.path.dirname(__file__))

    return os.path.join(base_path, relative_path)


# ============================================
# CARICAMENTO FIGURE DA CSV (con Area)
# ============================================

def load_figures(csv_path="costi_per_grade.csv"):
    """
    Ritorna:
    {
        "Cyber": {
            "Equity Partner": {"cost": ..., "bill_rate": ...},
            ...
        },
        "Specialist": { ... },
        ...
    }
    """
    figures_by_area = {}
    full_path = resource_path(csv_path)

    with open(full_path, newline="", encoding="utf-8") as f:
        # Il CSV è con separatore ';'
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            def conv(x):
                return float(x.replace(",", ".").strip())

            area = row["Area"].strip()
            role = row["Role"].strip()

            if area not in figures_by_area:
                figures_by_area[area] = {}

            figures_by_area[area][role] = {
                "cost": conv(row["Cost"]),
                "bill_rate": conv(row["Bill_Rate"]),
            }
    return figures_by_area


# ============================================
# FUNZIONE LP
# ============================================

def run_optimization(figures, min_percents, max_percents, gm_min, discount_max, project_value):
    recovery = 0.01
    HOURS_PER_DAY = 8

    roles = list(figures.keys())  # qui role_key = (area, role)

    model = pulp.LpProblem("Team_Optimization_Slack", pulp.LpMinimize)

    # k = numero di giornate intere; hours = 8 * k  → ore sempre multipli di 8
    k = pulp.LpVariable.dicts("days", roles, lowBound=0, cat="Integer")
    hours = {r: HOURS_PER_DAY * k[r] for r in roles}

    T = pulp.LpVariable("Total_Hours", lowBound=0)
    GM_diff = pulp.LpVariable("GM_diff", lowBound=0)
    Fee_diff = pulp.LpVariable("Fee_diff", lowBound=0)

    # totale ore
    model += T == pulp.lpSum(hours[r] for r in roles)

    # vincoli min% e max% per figura (sulle ore, che sono 8*k)
    for r in roles:
        model += hours[r] >= min_percents[r] * T
        model += hours[r] <= max_percents[r] * T

    # fee e costi
    GrossFees = pulp.lpSum(hours[r] * figures[r]["bill_rate"] for r in roles)
    TotalCost = pulp.lpSum(hours[r] * figures[r]["cost"] for r in roles)
    RecoveryAlloc = GrossFees * recovery
    GrossFeesTot = GrossFees + RecoveryAlloc

    GM = project_value - TotalCost
    GM_required = gm_min * project_value
    max_fees = project_value / (1 - discount_max) if (1 - discount_max) != 0 else project_value

    # slack GM
    model += GM_diff >= GM_required - GM
    model += GM_diff >= GM - GM_required

    # slack fee / discount
    model += Fee_diff >= max_fees - GrossFeesTot
    model += Fee_diff >= GrossFeesTot - max_fees

    # obiettivo
    model += 1000 * (GM_diff + Fee_diff) - T

    # 🛠️ fix permessi CBC prima di lanciare il solver
    ensure_cbc_executable()

    solver = pulp.PULP_CBC_CMD(msg=False)
    model.solve(solver)

    # 🔔 Check stato solver
    status = pulp.LpStatus[model.status]
    if status != "Optimal":
        raise RuntimeError(f"Soluzione ottima non trovata. Stato solver: {status}")

    # estrazione valori
    k_val = {r: k[r].value() for r in roles}
    hours_val = {r: (k_val[r] or 0) * HOURS_PER_DAY for r in roles}

    gross_fees = sum(hours_val[r] * figures[r]["bill_rate"] for r in roles)
    recovery_alloc = gross_fees * recovery
    gross_fees_tot = gross_fees + recovery_alloc
    total_cost = sum(hours_val[r] * figures[r]["cost"] for r in roles)

    GM_value = project_value - total_cost
    GM_pct = GM_value / project_value if project_value else 0
    discount_abs = gross_fees_tot - project_value
    discount_eff = discount_abs / gross_fees_tot if gross_fees_tot else 0

    total_days = sum(h / HOURS_PER_DAY for h in hours_val.values()) if HOURS_PER_DAY else 0
    daily_fee = project_value / total_days if total_days else 0

    return {
        "hours": hours_val,           # key = (area, role)
        "total_hours": T.value(),
        "total_days": total_days,
        "total_cost": total_cost,
        "project_value": project_value,
        "gross_fees": gross_fees,
        "recovery_allocation": recovery_alloc,
        "gross_fees_tot": gross_fees_tot,
        "GM_value": GM_value,
        "GM_pct": GM_pct,
        "GM_target": gm_min,
        "discount_abs": discount_abs,
        "discount_eff": discount_eff,
        "discount_target": discount_max,
        "daily_fee": daily_fee
    }


# worker per multiprocessing (deve stare al top-level)
def optimization_worker(figures_selected, min_percents, max_percents,
                        gm_val, discount_val, project_value, queue):
    try:
        result = run_optimization(
            figures_selected, min_percents, max_percents,
            gm_val, discount_val, project_value
        )
        queue.put(("ok", result))
    except Exception as e:
        queue.put(("err", str(e)))


# ============================================
# UI PRINCIPALE
# ============================================

def start_ui():

    figures_all = load_figures()  # dict[area][role]

    root = tk.Tk()
    root.title("Team Optimizer")
    # Nessuna geometry fissa: la calcoliamo dopo aver creato la UI

    # --------- STILE BOTTONI MINIMALE / MORBIDO ----------
    style = ttk.Style(root)
    style.configure(
        "Rounded.TButton",
        padding=(14, 6),
        relief="flat",
        borderwidth=1,
        font=("Segoe UI", 9),
    )
    style.map(
        "Rounded.TButton",
        relief=[("pressed", "sunken"), ("!pressed", "flat")]
    )

    # Funzione per adattare la finestra al contenuto
    def adjust_window_to_content():
        root.update_idletasks()
        req_w = root.winfo_reqwidth()
        req_h = root.winfo_reqheight()
        screen_w = root.winfo_screenwidth()
        screen_h = root.winfo_screenheight()

        # Limito a una percentuale dello schermo per non sforare
        width = min(req_w, int(screen_w * 0.95))
        height = min(req_h, int(screen_h * 0.9))

        root.geometry(f"{width}x{height}")
        root.minsize(width, height)

    # Layout principale
    main_frame = tk.Frame(root)
    main_frame.pack(fill="both", expand=True)

    main_frame.columnconfigure(0, weight=2)
    main_frame.columnconfigure(1, weight=10)
    main_frame.rowconfigure(0, weight=1)

    left_frame = tk.Frame(main_frame)
    left_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

    right_frame = tk.Frame(main_frame, bd=1, relief="groove", padx=10, pady=10)
    right_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

    # ===== INPUTS =====

    inputs_frame = tk.Frame(left_frame)
    inputs_frame.pack(fill="x", pady=5)

    # riga 0: Valore Progetto
    tk.Label(inputs_frame, text="Valore Progetto").grid(row=0, column=0, sticky="w", pady=2)
    proj_entry = tk.Entry(inputs_frame, width=10)
    proj_entry.grid(row=0, column=1, padx=(5, 2), pady=2)
    tk.Label(inputs_frame, text="€").grid(row=0, column=2, sticky="w", pady=2)

    # riga 1: GM %
    tk.Label(inputs_frame, text="GM % target").grid(row=1, column=0, sticky="w", pady=2)
    gm_entry = tk.Entry(inputs_frame, width=10)
    gm_entry.grid(row=1, column=1, padx=(5, 2), pady=2)
    tk.Label(inputs_frame, text="%").grid(row=1, column=2, sticky="w", pady=2)

    # riga 2: Discount %
    tk.Label(inputs_frame, text="Discount % massimo").grid(row=2, column=0, sticky="w", pady=2)
    discount_entry = tk.Entry(inputs_frame, width=10)
    discount_entry.grid(row=2, column=1, padx=(5, 2), pady=2)
    tk.Label(inputs_frame, text="%").grid(row=2, column=2, sticky="w", pady=2)

    inputs_frame.grid_columnconfigure(0, weight=1)

    tk.Label(
        left_frame,
        text="Seleziona figure (per Area) e partecipazione minima / massima (%):",
        font=("Arial", 10, "bold")
    ).pack(anchor="w", pady=(10, 5))

    # ===== NOTEBOOK A TAB PER AREA =====
    notebook = ttk.Notebook(left_frame)
    notebook.pack(fill="both", expand=True)

    checks = {}             # key = (area, role) -> BooleanVar
    minpercent_entries = {} # key = (area, role) -> Entry
    maxpercent_entries = {} # key = (area, role) -> Entry

    for area, roles_dict in figures_all.items():
        tab = tk.Frame(notebook)
        notebook.add(tab, text=area)

        container = tk.Frame(tab)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container)
        canvas.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")

        canvas.configure(yscrollcommand=scrollbar.set)

        scroll_frame = tk.Frame(canvas)
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")

        def on_configure(event, c=canvas):
            c.configure(scrollregion=c.bbox("all"))

        scroll_frame.bind("<Configure>", on_configure)

        # Mouse wheel scroll per tab
        def _on_mousewheel(event, c=canvas):
            c.yview_scroll(int(-event.delta / 120), "units")

        scroll_frame.bind(
            "<Enter>",
            lambda e, c=canvas: c.bind_all(
                "<MouseWheel>",
                lambda ev: _on_mousewheel(ev, c)
            ),
        )
        scroll_frame.bind(
            "<Leave>",
            lambda e, c=canvas: c.unbind_all("<MouseWheel>")
        )

        # Header riga
        tk.Label(scroll_frame, text="Figura").grid(row=0, column=0, sticky="w", pady=2, padx=(0, 10))
        tk.Label(scroll_frame, text="Min %").grid(row=0, column=1, sticky="w", pady=2)
        tk.Label(scroll_frame, text="").grid(row=0, column=2, sticky="w", pady=2)
        tk.Label(scroll_frame, text="Max %").grid(row=0, column=3, sticky="w", pady=2)
        tk.Label(scroll_frame, text="").grid(row=0, column=4, sticky="w", pady=2)

        # Righe per ruolo di quell'area
        for i, role in enumerate(roles_dict):
            row = i + 1
            key = (area, role)

            var = tk.BooleanVar(value=False)
            chk = tk.Checkbutton(scroll_frame, text=role, variable=var)
            chk.grid(row=row, column=0, sticky="w", pady=3, padx=(0, 10))

            # Min %
            min_entry = tk.Entry(scroll_frame, width=7)
            min_entry.grid(row=row, column=1, sticky="w", pady=3)
            min_entry.insert(0, '1')  # default 1%
            tk.Label(scroll_frame, text="%").grid(row=row, column=2, sticky="w", pady=3)

            # Max %
            max_entry = tk.Entry(scroll_frame, width=7)
            max_entry.grid(row=row, column=3, sticky="w", pady=3, padx=(10, 0))
            max_entry.insert(0, '60')  # default 60%
            tk.Label(scroll_frame, text="%").grid(row=row, column=4, sticky="w", pady=3)

            checks[key] = var
            minpercent_entries[key] = min_entry
            maxpercent_entries[key] = max_entry

        scroll_frame.grid_columnconfigure(0, weight=1)

    # ===== RISULTATI (Right Panel) =====

    tk.Label(
        right_frame,
        text="Composizione Team",
        font=("Arial", 12, "bold")
    ).pack(anchor="w", pady=(0, 10))

    # Frame + scrollbar per il Treeview
    tree_frame = tk.Frame(right_frame)
    tree_frame.pack(fill="both", expand=True, pady=(0, 10))

    tree = ttk.Treeview(tree_frame, columns=("role", "hours"), show="headings")
    tree.heading("role", text="Figura")
    tree.heading("hours", text="Ore")
    tree.column("role", width=260, anchor="w")
    tree.column("hours", width=80, anchor="center")
    tree.pack(side="left", fill="both", expand=True)

    tree_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    tree_scrollbar.pack(side="right", fill="y")

    tree.configure(yscrollcommand=tree_scrollbar.set)

    # Riga placeholder iniziale
    tree.insert("", "end", values=("Nessun risultato", "            "))

    # Riepilogo stile tabella
    info_frame = tk.Frame(right_frame)
    info_frame.pack(fill="x", pady=(5, 0))

    info_frame.grid_columnconfigure(0, weight=0)
    info_frame.grid_columnconfigure(1, weight=0)
    info_frame.grid_columnconfigure(2, weight=1)  # spacer
    info_frame.grid_columnconfigure(3, weight=0)
    info_frame.grid_columnconfigure(4, weight=0)

    def new_var():
        # placeholder "lungo" così occupa più o meno lo spazio dei numeri veri
        return tk.StringVar(value="                                     ")

    total_left_var = new_var()
    grossfees_right_var = new_var()
    gmperc_left_var = new_var()
    recalloc_right_var = new_var()
    gm_left_var = new_var()
    tot_right_var = new_var()
    dailyfee_left_var = new_var()
    totadj_right_var = new_var()
    totalfee_left_var = new_var()
    discount_right_var = new_var()

    def add_info_row(row, left_label, left_var, right_label, right_var):
        # blocco sinistro
        tk.Label(info_frame, text=left_label).grid(row=row, column=0, sticky="w", pady=2)
        tk.Label(info_frame, textvariable=left_var, font=("Arial", 9, "bold")).grid(
            row=row, column=1, sticky="w", pady=2, padx=(5, 10)
        )
        # blocco destro
        tk.Label(info_frame, text=right_label).grid(
            row=row, column=3, sticky="e", pady=2, padx=(10, 5)
        )
        tk.Label(info_frame, textvariable=right_var, font=("Arial", 9, "bold")).grid(
            row=row, column=4, sticky="e", pady=2
        )

    add_info_row(0, "Total (gg/h/costo):", total_left_var, "Gross Fees:", grossfees_right_var)
    add_info_row(1, "GM %:", gmperc_left_var, "Recovery Allocation (1%):", recalloc_right_var)
    add_info_row(2, "GM:", gm_left_var, "TOT:", tot_right_var)
    add_info_row(3, "Daily Fee:", dailyfee_left_var, "TotAdj:", totadj_right_var)
    add_info_row(4, "Total Fee:", totalfee_left_var, "Discount:", discount_right_var)

    last_result = {"data": None}

    def update_results(result):

        for item in tree.get_children():
            tree.delete(item)

        for key, h in result["hours"].items():
            area, role = key
            label = f"{area} - {role}"
            tree.insert("", "end", values=(label, f"{h:.0f}"))

        days = result["total_days"]
        hours_tot = result["total_hours"]
        cost_tot = result["total_cost"]
        gm_val = result["GM_value"]
        gm_pct = result["GM_pct"]
        gross_fees = result["gross_fees"]
        rec_alloc = result["recovery_allocation"]
        gross_fees_tot = result["gross_fees_tot"]
        daily_fee = result["daily_fee"]
        project_value = result["project_value"]
        discount_abs = result["discount_abs"]
        discount_eff = result["discount_eff"]

        total_left_var.set(f"{days:.0f} gg | {hours_tot:.0f} h | {euro_fmt(cost_tot)}")
        grossfees_right_var.set(euro_fmt(gross_fees))
        gmperc_left_var.set(pct_fmt(gm_pct))
        recalloc_right_var.set(euro_fmt(rec_alloc))
        gm_left_var.set(euro_fmt(gm_val))
        tot_right_var.set(euro_fmt(gross_fees_tot))
        dailyfee_left_var.set(euro_fmt(daily_fee))
        totadj_right_var.set(euro_fmt(discount_abs))
        totalfee_left_var.set(euro_fmt(project_value))
        discount_right_var.set(pct_fmt(discount_eff))

        last_result["data"] = result

        # Dopo aver popolato i risultati, rivaluto lo spazio necessario
        adjust_window_to_content()

    # ===== EXPORT EXCEL =====

    def save_excel():
        result = last_result["data"]
        if result is None:
            messagebox.showwarning("Attenzione", "Esegui prima un calcolo.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")]
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # --- Ore per figura ---
        ws.append(["Area", "Role", "Hours"])
        for key, h in result["hours"].items():
            area, role = key
            ws.append([area, role, h])

        ws.append([])

        # --- Tabella riepilogo ---
        days = result["total_days"]
        hours_tot = result["total_hours"]
        cost_tot = result["total_cost"]
        gm_val = result["GM_value"]
        gm_pct = result["GM_pct"]
        gross_fees = result["gross_fees"]
        rec_alloc = result["recovery_allocation"]
        gross_fees_tot = result["gross_fees_tot"]
        daily_fee = result["daily_fee"]
        project_value = result["project_value"]
        discount_abs = result["discount_abs"]
        discount_eff = result["discount_eff"]

        ws.append(["Total", days, hours_tot, "Cost Tot", cost_tot])
        r1 = ws.max_row
        ws.append(["Gross Fees", gross_fees, "", "Recovery Alloc", rec_alloc])
        r2 = ws.max_row
        ws.append(["GM", gm_val, "GM %", gm_pct, "TOT", gross_fees_tot])
        r3 = ws.max_row
        ws.append(["Daily Fee", daily_fee, "", "TotAdj", discount_abs])
        r4 = ws.max_row
        ws.append(["Project Value", project_value, "", "Discount", discount_eff])
        r5 = ws.max_row

        money_cells = [
            f"E{r1}",
            f"B{r2}", f"E{r2}",
            f"B{r3}", f"E{r3}",
            f"B{r4}", f"E{r4}",
            f"B{r5}",
        ]
        for cell in money_cells:
            ws[cell].number_format = "#,##0.00"

        ws[f"D{r3}"].number_format = "0.00%"
        ws[f"E{r5}"].number_format = "0.00%"

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

        wb.save(path)
        messagebox.showinfo("Salvato", "Risultati salvati in Excel con successo.")

    # bottone export (minimal)
    excel_btn = ttk.Button(
        right_frame,
        text="Esporta in Excel",
        command=save_excel,
        style="Rounded.TButton"
    )
    excel_btn.pack(anchor="center", pady=(10, 0))

    # ===== STATO PROCESSO CORRENTE =====
    current_process = {"proc": None, "queue": None}

    # ===== SPINNER / LOADER AL POSTO DEL BOTTONE CALCOLA =====

    calc_area = ttk.Frame(left_frame)
    calc_area.pack(pady=10)

    calc_btn = ttk.Button(calc_area, text="Calcola", style="Rounded.TButton")
    calc_btn.pack()  # inizialmente visibile

    spinner = ttk.Progressbar(calc_area, mode="indeterminate", length=140)
    cancel_btn = ttk.Button(calc_area, text="✕", width=2, style="Rounded.TButton")

    def show_spinner():
        calc_btn.pack_forget()
        spinner.pack(side="left")
        cancel_btn.pack(side="left", padx=(8, 0))
        spinner.start(10)

    def hide_spinner():
        spinner.stop()
        spinner.pack_forget()
        cancel_btn.pack_forget()
        calc_btn.pack()

    # ===== CALLBACK FINE CALCOLO =====

    last_result = {"data": None}

    def on_calc_finished(result, error):
        hide_spinner()
        calc_btn.config(state="normal")
        excel_btn.config(state="normal")

        if error is not None:
            messagebox.showerror("Errore", f"Errore durante l'ottimizzazione:\n{error}")
            return

        update_results(result)

    # ===== CANCELLAZIONE CALCOLO =====

    def cancel_calc():
        proc = current_process.get("proc")
        if proc is not None and proc.is_alive():
            proc.terminate()
        current_process["proc"] = None
        current_process["queue"] = None
        hide_spinner()
        calc_btn.config(state="normal")
        excel_btn.config(state="normal")
        messagebox.showinfo("Annullato", "Calcolo annullato dall'utente.")

    cancel_btn.config(command=cancel_calc)

    # ===== CALCOLO =====

    def run():
        # recupero input
        try:
            project_value = float(proj_entry.get().replace(",", "."))
        except Exception:
            messagebox.showerror("Errore", "Valore progetto non valido.")
            return

        try:
            gm_val = parse_percent_no_symbol(gm_entry.get())
            discount_val = parse_percent_no_symbol(discount_entry.get())
        except Exception:
            messagebox.showerror("Errore", "Valori GM o Discount non validi.")
            return

        selected_roles = [key for key, v in checks.items() if v.get()]
        if not selected_roles:
            messagebox.showerror("Errore", "Seleziona almeno una figura.")
            return

        # Flatten figures selezionate: key = (area, role)
        figures_selected = {
            key: figures_all[key[0]][key[1]]
            for key in selected_roles
        }

        min_percents = {}
        max_percents = {}
        total_min_percent = 0

        for key in selected_roles:
            min_text = minpercent_entries[key].get().strip()
            max_text = maxpercent_entries[key].get().strip()

            # default: min vuota -> 0%, max vuota -> 100%
            if min_text == "":
                p_min = 0.0
            else:
                try:
                    p_min = parse_percent_no_symbol(min_text)
                except Exception:
                    messagebox.showerror("Errore", f"Min% non valido per {key[0]} - {key[1]}")
                    return

            if max_text == "":
                p_max = 1.0
            else:
                try:
                    p_max = parse_percent_no_symbol(max_text)
                except Exception:
                    messagebox.showerror("Errore", f"Max% non valido per {key[0]} - {key[1]}")
                    return

            # controlli
            if p_min < 0 or p_max < 0:
                messagebox.showerror("Errore", f"Percentuali negative non permesse ({key[0]} - {key[1]}).")
                return
            if p_max > 1.0:
                messagebox.showerror("Errore", f"Max% per {key[0]} - {key[1]} supera il 100%.")
                return
            if p_min > p_max:
                messagebox.showerror("Errore", f"Per {key[0]} - {key[1]} il Min% è maggiore del Max%.")
                return

            min_percents[key] = p_min
            max_percents[key] = p_max
            total_min_percent += p_min

        if total_min_percent > 1.0:
            messagebox.showerror(
                "Errore",
                f"La somma dei min% ({total_min_percent * 100:.2f}%) supera il 100%."
            )
            return

        # avvia il calcolo in background (processo separato)
        calc_btn.config(state="disabled")
        excel_btn.config(state="disabled")
        show_spinner()

        q = multiprocessing.Queue()
        proc = multiprocessing.Process(
            target=optimization_worker,
            args=(figures_selected, min_percents, max_percents,
                  gm_val, discount_val, project_value, q)
        )
        current_process["proc"] = proc
        current_process["queue"] = q
        proc.start()

        def poll_queue():
            proc_local = current_process.get("proc")
            q_local = current_process.get("queue")

            if proc_local is None or q_local is None:
                return  # annullato o già gestito

            if not q_local.empty():
                status, payload = q_local.get()
                current_process["proc"] = None
                current_process["queue"] = None
                if status == "ok":
                    on_calc_finished(payload, None)
                else:
                    on_calc_finished(None, RuntimeError(payload))
                return

            if not proc_local.is_alive():
                # processo morto senza mettere nulla in coda
                current_process["proc"] = None
                current_process["queue"] = None
                on_calc_finished(None, RuntimeError("Il solver si è interrotto senza restituire risultato."))
                return

            root.after(100, poll_queue)

        root.after(100, poll_queue)

    calc_btn.config(command=run)

    # ===== PRIMO ADATTAMENTO DELLA FINESTRA AL CONTENUTO DI BASE =====
    adjust_window_to_content()

    root.mainloop()


# ============================================
# MAIN
# ============================================

if __name__ == "__main__":
    multiprocessing.freeze_support()  # fondamentale per exe su Windows
    start_ui()
